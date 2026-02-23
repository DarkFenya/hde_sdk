import os
from datetime import datetime

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from clients.api_client import HdeApi

load_dotenv()

client = HdeApi(os.getenv("HDE_TOKEN"), os.getenv("HDE_EMAIL"), os.getenv("HDE_BASE_URL"))


def export_users_to_excel(filename: str = None):
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"users_export_{timestamp}.xlsx"

    print(f"Экспорт в {filename}...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    ws.append(["ID", "Email"])

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    total = 0
    for page_num, page in enumerate(client.users.get_users_lazy(), start=1):
        print(f"Страница {page_num}: {len(page)} пользователей")
        for user in page:
            ws.append([user.get("id", ""), user.get("email", "")])
            total += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 35
    wb.save(filename)

    print(f"Готово: {total} пользователей → {filename}")


if __name__ == "__main__":
    export_users_to_excel()
