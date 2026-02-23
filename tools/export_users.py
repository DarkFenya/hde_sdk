"""
Быстрый экспорт пользователей через asyncio.
Параллельно загружает несколько страниц — быстрее чем простой вариант.
"""
import asyncio
import os
from datetime import datetime

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from clients.api_client import HdeApi

load_dotenv()

client = HdeApi(os.getenv("HDE_TOKEN"), os.getenv("HDE_EMAIL"), os.getenv("HDE_BASE_URL"))

MAX_CONCURRENT_REQUESTS = 20


async def fetch_page(semaphore, page_num):
    async with semaphore:
        return await asyncio.to_thread(client.users.get_users_page, page=page_num)


async def fetch_all_users(max_concurrent: int = MAX_CONCURRENT_REQUESTS) -> list:
    first = await asyncio.to_thread(client.users.get_users_page, page=1)
    if not first:
        return []

    data = first.json()
    pagination = data.get("pagination", {})
    total_pages = pagination.get("total_pages", 1)
    print(f"Всего страниц: {total_pages}, пользователей: {pagination.get('total', '?')}")

    all_users = data.get("data", [])

    if total_pages > 1:
        semaphore = asyncio.Semaphore(max_concurrent)
        tasks = [fetch_page(semaphore, p) for p in range(2, total_pages + 1)]
        responses = await asyncio.gather(*tasks)
        for resp in responses:
            if resp:
                all_users.extend(resp.json().get("data", []))

    return all_users


async def export_users_to_excel(filename: str = None, max_concurrent: int = MAX_CONCURRENT_REQUESTS):
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"users_export_{timestamp}.xlsx"

    print(f"Экспорт в {filename} ({max_concurrent} параллельных запросов)...")
    users = await fetch_all_users(max_concurrent)

    if not users:
        print("Пользователей не найдено.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"
    ws.append(["ID", "Email"])

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for user in users:
        ws.append([user.get("id", ""), user.get("email", "")])

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 35

    await asyncio.to_thread(wb.save, filename)
    print(f"Готово: {len(users)} пользователей → {filename}")


if __name__ == "__main__":
    asyncio.run(export_users_to_excel())
